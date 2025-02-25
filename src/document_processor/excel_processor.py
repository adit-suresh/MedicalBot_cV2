import pandas as pd
import logging
from typing import Dict, List, Optional, Tuple
import os
from datetime import datetime
from openpyxl import load_workbook
import re

logger = logging.getLogger(__name__)

class EnhancedExcelProcessor:
    """Enhanced processor for handling Excel files with multiple rows and proper date formatting."""

    def __init__(self):
        """Initialize Excel processor."""
        # Define required fields and their types/formats
        self.required_fields = {
            'contract_name': str,
            'first_name': str,
            'middle_name': str,
            'last_name': str,
            'effective_date': 'date',
            'dob': 'date',
            'gender': ['Male', 'Female'],
            'marital_status': ['Married', 'Divorced', 'Widowed', 'Single'],
            'category': str,
            'relation': str,
            'principal_card_no': str,
            'family_no': str,
            'staff_id': str,
            'nationality': str,
            'emirates_id': str,
            'unified_no': str,
            'passport_no': str,
            'work_country': str,
            'work_emirate': str,
            'work_region': str,
            'residence_country': str,
            'residence_emirate': str,
            'residence_region': str,
            'email': str,
            'mobile_no': str,
            'salary_band': str,
            'commission': str,
            'visa_issuance_emirate': str,
            'visa_file_number': str,
            'member_type': str
        }

        # Default value for missing fields
        self.DEFAULT_VALUE = "."
        
        # Date format for standardization
        self.DATE_FORMAT = '%d-%m-%Y'  # Changed to dd-mm-yyyy
        
    def validate_template(self, file_path: str) -> Tuple[bool, List[str]]:
        """
        Validate Excel template against required fields.
        
        Args:
            file_path: Path to Excel template file
            
        Returns:
            Tuple of (validity, missing fields)
        """
        try:
            # Load Excel file
            wb = load_workbook(file_path)
            ws = wb.active

            # Get header row
            header = [cell.value for cell in ws[1]]
            missing_fields = [field for field in self.required_fields.keys() if field not in header]

            return len(missing_fields) == 0, missing_fields

        except Exception as e:
            logger.error(f"Error validating Excel template: {str(e)}")
            raise Exception(f"Excel template validation failed: {str(e)}")
        
    def process_excel(self, file_path: str, dayfirst: bool = True) -> Tuple[pd.DataFrame, List[Dict]]:
        """
        Process Excel file and validate data. Handles multiple rows.
        
        Args:
            file_path: Path to Excel file
            dayfirst: Whether to parse dates with day first (European format)
            
        Returns:
            Tuple of (processed dataframe, list of validation errors)
        """
        try:
            df = pd.read_excel(file_path)
            
            # If empty, return empty dataframe
            if df.empty:
                logger.warning("Excel file contains no data")
                return df, [{"error": "Excel file contains no data"}]
            
            # Clean column names
            df.columns = [self._clean_column_name(col) for col in df.columns]

            # Fill missing values with default
            for field in self.required_fields.keys():
                if field not in df.columns:
                    df[field] = self.DEFAULT_VALUE

            # Validate data
            errors = self._validate_data(df)

            # Clean and standardize data
            df = self._clean_data(df, dayfirst=dayfirst)

            return df, errors

        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            raise Exception(f"Excel processing failed: {str(e)}")

    def _clean_column_name(self, column: str) -> str:
        """Clean and standardize column names."""
        if not isinstance(column, str):
            return ''
        # Remove special characters and spaces
        clean_name = re.sub(r'[^a-zA-Z0-9_]', '_', column.lower().strip())
        # Remove multiple underscores
        clean_name = re.sub(r'_+', '_', clean_name)
        # Remove leading/trailing underscores
        return clean_name.strip('_')

    def _validate_data(self, df: pd.DataFrame) -> List[Dict]:
        """Validate DataFrame against required fields and formats."""
        errors = []

        for field, validation in self.required_fields.items():
            if field not in df.columns:
                errors.append({
                    'field': field,
                    'error': 'Missing required column'
                })
                continue

            # Validate based on type
            if validation == 'date':
                try:
                    pd.to_datetime(df[field], errors='raise')
                except Exception as e:
                    errors.append({
                        'field': field,
                        'error': f'Invalid date format: {str(e)}'
                    })
            elif isinstance(validation, list):
                # Only check non-empty values
                invalid_mask = ~df[field].isin(validation) & ~df[field].isin(['', '.', self.DEFAULT_VALUE])
                invalid_values = df.loc[invalid_mask, field].unique()
                if len(invalid_values) > 0:
                    errors.append({
                        'field': field,
                        'error': f'Invalid values: {invalid_values}',
                        'valid_values': validation
                    })

        return errors

    def _clean_data(self, df: pd.DataFrame, dayfirst: bool = True) -> pd.DataFrame:
        """Clean and standardize data with improved date handling."""
        # Handle dates
        date_fields = ['dob', 'effective_date', 'date_of_birth']
        
        for field in date_fields:
            if field in df.columns:
                try:
                    # Convert to datetime and then to the desired format
                    df[field] = pd.to_datetime(df[field], dayfirst=dayfirst).dt.strftime('%d-%m-%Y')
                    # Format to DD-MM-YYYY
                    df[field] = df[field].dt.strftime(self.DATE_FORMAT)
                except:
                    df[field] = self.DEFAULT_VALUE

        # Clean string fields
        string_fields = [field for field, val_type in self.required_fields.items() 
                        if val_type == str and field in df.columns]
        for field in string_fields:
            df[field] = df[field].fillna(self.DEFAULT_VALUE)
            df[field] = df[field].astype(str).str.strip()

        # Handle phone numbers
        if 'mobile_no' in df.columns:
            df['mobile_no'] = df['mobile_no'].apply(self._format_phone_number)

        # Handle email addresses
        if 'email' in df.columns:
            df['email'] = df['email'].str.lower()

        return df

    def _format_phone_number(self, number: str) -> str:
        """Format phone numbers consistently."""
        if pd.isna(number) or number == self.DEFAULT_VALUE:
            return self.DEFAULT_VALUE
        
        # Remove non-numeric characters
        number = re.sub(r'\D', '', str(number))
        
        # Add country code if missing
        if len(number) == 9:  # Local number without country code
            number = '971' + number
        elif len(number) == 10 and number.startswith('0'):  # Local number with leading 0
            number = '971' + number[1:]
            
        return f"+{number}"
        
    def populate_template(self, template_path: str, output_path: str, data: List[Dict]) -> Dict:
        """
        Populate an Excel template with multiple rows of data.
        
        Args:
            template_path: Path to template Excel
            output_path: Path to save populated Excel
            data: List of dictionaries with data to populate
            
        Returns:
            Dict with status and info
        """
        try:
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template file not found: {template_path}")
                
            # Load template
            template_wb = load_workbook(template_path)
            template_ws = template_wb.active
            
            # Get header row with cleaned names
            headers = [self._clean_column_name(cell.value or '') for cell in template_ws[1]]
            
            # Create mapping from clean names to column indices
            header_map = {name: idx+1 for idx, name in enumerate(headers) if name}
            
            # Create new workbook for output
            output_wb = load_workbook(template_path)
            output_ws = output_wb.active
            
            # Populate data rows
            for row_idx, row_data in enumerate(data, start=2):  # Start from second row
                for field, value in row_data.items():
                    clean_field = self._clean_column_name(field)
                    if clean_field in header_map:
                        col_idx = header_map[clean_field]
                        output_ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save output
            output_wb.save(output_path)
            
            return {
                "status": "success",
                "template": template_path,
                "output": output_path,
                "rows_processed": len(data)
            }
            
        except Exception as e:
            logger.error(f"Error populating template: {str(e)}")
            return {
                "status": "error",
                "error": str(e)
            }
        
        
    def _process_date_field(self, df: pd.DataFrame, field: str) -> pd.DataFrame:
        """Process date field ensuring DD/MM/YYYY format."""
        try:
            if field in df.columns:
                # Convert to datetime with dayfirst=True
                df[field] = pd.to_datetime(df[field], dayfirst=True)
                # Format as DD/MM/YYYY
                df[field] = df[field].dt.strftime('%d/%m/%Y')
        except Exception as e:
            logger.error(f"Error processing date field {field}: {e}")
        return df

    def process_excel(self, file_path: str, dayfirst: bool = True) -> Tuple[pd.DataFrame, List[Dict]]:
        """Process Excel file with proper date handling."""
        try:
            df = pd.read_excel(file_path)
            if df.empty:
                return df, []
                
            # Handle date fields properly
            date_fields = ['DOB', 'Effective Date', 'dob', 'effective_date', 
                        'Passport Expiry Date', 'Visa Expiry Date']
            
            for field in date_fields:
                df = self._process_date_field(df, field)
            
            # Initialize errors list
            errors = []
            
            # Rest of your existing processing code...
            
            return df, errors
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            raise Exception(f"Excel processing failed: {str(e)}")